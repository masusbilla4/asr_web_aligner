# -*- coding: utf-8 -*-
"""Flask web app for Filipino ASR Evaluation - AI Aligner with Sentence Evaluation."""
import os
import io
import json
import re
import math
import time
import random
import datetime
import threading
from flask import Flask, render_template, request, jsonify, send_file, make_response

from alignment_engine import run_alignment, align_translation_local

app = Flask(__name__)

# --- Configuration ---
DEFAULT_MODEL = "gemini-2.5-flash"
AVAILABLE_MODELS = [
    "gemini-2.5-flash", "gemini-2.5-flash-lite", "gemini-2.5-pro",
    "gemini-2.0-flash", "gemini-1.5-flash", "gemini-1.5-pro"
]
MODEL_RPM_LIMITS = {
    "gemini-2.5-flash": 15, "gemini-2.5-flash-lite": 15,
    "gemini-2.5-pro": 10, "gemini-2.0-flash": 15,
    "gemini-1.5-flash": 15, "gemini-1.5-pro": 10
}
BATCH_SIZE = 40
OVERLAP_CONTEXT = 3
RETRY_DELAY = 10
MAX_RETRIES = 5
BATCH_DELAY = 15

EXPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)

# --- Real-time progress tracking ---
_progress_store = {
    "logs": [],
    "progress": 0,
    "status": "",
    "running": False,
    "task_id": 0,
    "partial_data": None,
    "partial_count": 0
}

def add_log(msg):
    _progress_store["logs"].append(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}")
    if len(_progress_store["logs"]) > 500:
        _progress_store["logs"] = _progress_store["logs"][-500:]

def reset_progress(task_name=""):
    _progress_store["task_id"] += 1
    _progress_store["logs"] = []
    _progress_store["progress"] = 0
    _progress_store["status"] = task_name
    _progress_store["running"] = True
    if task_name:
        add_log(f"▶ Started: {task_name}")

def finish_progress(msg="Done"):
    _progress_store["running"] = False
    _progress_store["progress"] = 100
    add_log(f"✅ {msg}")


@app.route("/api/progress")
def api_progress():
    return jsonify({
        "logs": _progress_store["logs"],
        "progress": _progress_store["progress"],
        "status": _progress_store["status"],
        "running": _progress_store["running"],
        "task_id": _progress_store["task_id"],
        "partial_count": _progress_store["partial_count"]
    })


@app.route("/api/cancel", methods=["POST"])
def api_cancel():
    """Cancel the running AI task."""
    if _progress_store["running"]:
        _progress_store["running"] = False
        add_log("🛑 Task cancelled by user")
        _progress_store["status"] = "Cancelled"
        return jsonify({"status": "cancelled", "partial_count": _progress_store["partial_count"]})
    return jsonify({"status": "not_running"})


@app.route("/api/partial-results")
def api_partial_results():
    """Fetch partial AI results (available even if task is still running or was cancelled)."""
    data = _progress_store.get("partial_data")
    if data is None:
        return jsonify({"error": "No partial results available."}), 404
    return jsonify({
        "alignment_data": data,
        "count": _progress_store["partial_count"],
        "running": _progress_store["running"]
    })


# --- Routes ---
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/align", methods=["POST"])
def api_align():
    data = request.json
    true_text = data.get("true_text", "").strip()
    asr_text = data.get("asr_text", "").strip()

    if not true_text or not asr_text:
        return jsonify({"error": "Both True Text and ASR Result are required."}), 400

    true_lines = [l.strip() for l in true_text.split("\n") if l.strip()]
    asr_lines = [l.rstrip() for l in asr_text.splitlines() if l.strip()]

    # Word count mismatch warning
    ref_wc = sum(len(l.split()) for l in true_lines)
    hyp_wc = sum(len(l.split()) for l in asr_lines)
    if ref_wc > 0 and hyp_wc > 0:
        ratio = max(ref_wc, hyp_wc) / min(ref_wc, hyp_wc)
        if ratio > 3:
            return jsonify({
                "warning": True,
                "message": f"Significant word count mismatch: True Text={ref_wc}, ASR={hyp_wc}, Ratio={ratio:.1f}x. Results may be inaccurate."
            })

    try:
        result = run_alignment(true_lines, asr_lines)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/align-force", methods=["POST"])
def api_align_force():
    """Force alignment even with mismatch warning."""
    data = request.json
    true_text = data.get("true_text", "").strip()
    asr_text = data.get("asr_text", "").strip()

    true_lines = [l.strip() for l in true_text.split("\n") if l.strip()]
    asr_lines = [l.rstrip() for l in asr_text.splitlines() if l.strip()]

    try:
        result = run_alignment(true_lines, asr_lines)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/align-translation-local", methods=["POST"])
def api_align_translation_local():
    data = request.json
    trans_blob = data.get("translation_text", "").strip()
    alignment_data = data.get("alignment_data", [])

    if not trans_blob:
        return jsonify({"error": "Translation text is required."}), 400
    if not alignment_data:
        return jsonify({"error": "Run ASR alignment first."}), 400

    translations, method = align_translation_local(alignment_data, trans_blob)

    for i, trans in enumerate(translations):
        if i < len(alignment_data):
            alignment_data[i]["translation"] = trans

    return jsonify({
        "alignment_data": alignment_data,
        "method": method,
        "count": len(translations)
    })


@app.route("/api/align-translation-ai", methods=["POST"])
def api_align_translation_ai():
    data = request.json
    trans_blob = data.get("translation_text", "").strip()
    alignment_data = data.get("alignment_data", [])
    api_key = data.get("api_key", "") or os.getenv("GEMINI_API_KEY", "")
    model = data.get("model", DEFAULT_MODEL)

    if not trans_blob:
        return jsonify({"error": "Translation text is required."}), 400
    if not alignment_data:
        return jsonify({"error": "Run ASR alignment first."}), 400
    if not api_key:
        return jsonify({"error": "Gemini API key is required."}), 400

    if _progress_store["running"]:
        return jsonify({"error": "Another AI task is already running. Wait for it to finish or check partial results."}), 400

    # Set running=True HERE so the first poll sees it
    _progress_store["running"] = True
    _progress_store["status"] = "Starting AI Translation Alignment..."
    _progress_store["progress"] = 0
    _progress_store["logs"] = []
    _progress_store["partial_data"] = None
    _progress_store["partial_count"] = 0
    add_log("▶ Starting AI Translation Alignment...")

    # Run in background thread
    task_id = _progress_store["task_id"] + 1
    t = threading.Thread(target=_run_translation_ai, args=(alignment_data, trans_blob, api_key, model, task_id), daemon=True)
    t.start()
    return jsonify({"status": "started", "task_id": task_id})


def _run_translation_ai(alignment_data, trans_blob, api_key, model, task_id):
    try:
        from google import genai
        from google.genai import types
        client = genai.Client(api_key=api_key)
    except ImportError:
        add_log("❌ google-genai not installed")
        finish_progress("Failed: google-genai not installed")
        return
    except Exception as e:
        add_log(f"❌ API init failed: {str(e)}")
        finish_progress(f"Failed: {str(e)}")
        return

    total_sentences = len(alignment_data)
    all_translations = {}
    total_batches = math.ceil(total_sentences / BATCH_SIZE)

    # Don't call reset_progress here - state already set in route handler
    _progress_store["status"] = f"AI Translation Alignment ({total_sentences} sentences, {total_batches} batches)"
    add_log(f"Model: {model} | Batch size: {BATCH_SIZE} | Overlap: {OVERLAP_CONTEXT}")

    batch_num = 0
    batch_start = 0

    while batch_start < total_sentences:
        if not _progress_store["running"]:
            add_log("🛑 Task stopped by user")
            break
        batch_end = min(batch_start + BATCH_SIZE, total_sentences)
        context_start = max(0, batch_start - OVERLAP_CONTEXT)
        context_end = min(total_sentences, batch_end + OVERLAP_CONTEXT)
        core_start = batch_start
        core_end = batch_end

        asr_sentences_with_context = []
        for i in range(context_start, context_end):
            d = alignment_data[i]
            is_core = core_start <= i < core_end
            is_context_before = i < core_start
            is_context_after = i >= core_end
            sentence_info = {
                "id": d["id"],
                "asr_text": d["asr_nobreak"],
                "is_core": is_core,
                "is_context_before": is_context_before,
                "is_context_after": is_context_after
            }
            if is_context_before and i in all_translations:
                sentence_info["previous_translation"] = all_translations[i]
            asr_sentences_with_context.append(sentence_info)

        context_info = ""
        if context_start < core_start:
            context_info += f"\nCONTEXT BEFORE (sentences {context_start+1}-{core_start}): For context.\n"
        if context_end > core_end:
            context_info += f"\nCONTEXT AFTER (sentences {batch_end+1}-{context_end}): For context.\n"

        core_count = core_end - core_start

        prompt = f"""TASK: Separate the English translation blob to match each ASR Result segment.

BATCH INFO:
- Batch {batch_num+1} of {math.ceil(total_sentences/BATCH_SIZE)}
- Core sentences: {core_start+1} to {core_end} ({core_count} sentences)
{context_info}

ASR SEGMENTS:
{json.dumps(asr_sentences_with_context, indent=2, ensure_ascii=False)}

ENGLISH TRANSLATION BLOB:
{trans_blob}

INSTRUCTIONS:
1. Return exactly {core_count} translation segments for CORE sentences only.
2. Each translation corresponds to ONE ASR segment by ID.
3. Match MEANING, ignore punctuation differences.

OUTPUT FORMAT (JSON):
{{
  "translations": ["translation 1", "translation 2", ...],
  "first_sentence_id": {core_start+1},
  "last_sentence_id": {core_end}
}}

The translations array must have exactly {core_count} elements."""

        add_log(f"📦 Batch {batch_num+1}/{total_batches}: sentences {core_start+1}-{core_end} ({core_count} sentences)")
        _progress_store["progress"] = int((batch_num / total_batches) * 100)

        for retry_attempt in range(MAX_RETRIES):
            try:
                add_log(f"  📡 Sending API request (attempt {retry_attempt+1})...")
                response = client.models.generate_content(
                    model=model, contents=prompt,
                    config=types.GenerateContentConfig(response_mime_type='application/json')
                )
                if response.text:
                    response_text = response.text.strip()
                    if response_text.startswith('```'):
                        response_text = re.sub(r'^```json?\s*', '', response_text)
                        response_text = re.sub(r'\s*```$', '', response_text)
                    result = json.loads(response_text)
                    translations = result.get("translations", [])
                    first_id = result.get("first_sentence_id", core_start + 1)
                    matched = 0
                    for i, trans in enumerate(translations):
                        sent_idx = first_id - 1 + i
                        if core_start <= sent_idx < core_end:
                            all_translations[sent_idx] = trans
                            matched += 1
                    add_log(f"  ✅ Got {matched} translations (API returned {len(translations)})")
                    break
            except Exception as e:
                add_log(f"  ⚠️ Retry {retry_attempt+1}: {str(e)[:80]}")
                if retry_attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY * (2 ** retry_attempt))

        # Store partial results after each batch
        for i in range(len(alignment_data)):
            if i in all_translations:
                alignment_data[i]["translation"] = all_translations[i]
        _progress_store["partial_data"] = alignment_data
        _progress_store["partial_count"] = len(all_translations)
        add_log(f"  💾 Partial results saved: {len(all_translations)}/{total_sentences} sentences")

        batch_start = batch_end
        batch_num += 1
        if batch_start < total_sentences:
            add_log(f"  ⏳ Waiting {BATCH_DELAY}s before next batch...")
            time.sleep(BATCH_DELAY)

    finish_progress(f"Translation aligned: {len(all_translations)} sentences")



@app.route("/api/evaluate-ai", methods=["POST"])
def api_evaluate_ai():
    data = request.json
    alignment_data = data.get("alignment_data", [])
    api_key = data.get("api_key", "") or os.getenv("GEMINI_API_KEY", "")
    model = data.get("model", DEFAULT_MODEL)

    if not alignment_data:
        return jsonify({"error": "Run alignment first."}), 400
    if not api_key:
        return jsonify({"error": "Gemini API key is required."}), 400

    try:
        from google import genai
        from google.genai import types
        client = genai.Client(api_key=api_key)
    except ImportError:
        return jsonify({"error": "Install google-genai: pip install google-genai"}), 400
    except Exception as e:
        return jsonify({"error": f"API init failed: {str(e)}"}), 400

    total_sentences = len(alignment_data)
    all_evaluations = {}
    total_batches = math.ceil(total_sentences / BATCH_SIZE)

    reset_progress(f"AI Evaluation ({total_sentences} sentences, {total_batches} batches)")
    add_log(f"Model: {model} | Batch size: {BATCH_SIZE}")

    for batch_idx, batch_start in enumerate(range(0, total_sentences, BATCH_SIZE)):
        batch_end = min(batch_start + BATCH_SIZE, total_sentences)
        batch_data = alignment_data[batch_start:batch_end]

        sentences = []
        for d in batch_data:
            sentences.append({"id": d["id"], "true": d["true"], "asr": d["asr_nobreak"]})

        prompt = f"""TASK: Evaluate ASR transcription quality for each sentence pair.

SCORING GUIDE:
- 3: No wrong/missing/additional word. Perfect sentence.
- 2.5: Some errors but no problem understanding the sentence.
- 2: Some errors + grammar issues, but overall meaning is clear.
- 1.5: Partially able to guess and understand subject/details.
- 1: Unable to understand or no text transcribed.

SPECIAL RULES:
- Missing punctuation causing meaning confusion → score 1 or 1.5
- Evaluate per sentence, not per dialogue

INPUT SENTENCES (JSON):
{json.dumps(sentences, indent=2, ensure_ascii=False)}

OUTPUT FORMAT (JSON only):
{{
  "evaluations": [
    {{"id": <sentence_id>, "score": <1|1.5|2|2.5|3>, "reason": "<brief reason>"}},
    ...
  ]
}}

REASON EXAMPLES: "Perfect match", "Missing punctuation", "Wrong word detection", "Missing word", "Additional word", "Grammar error", "Unintelligible"

Return ONLY valid JSON with exactly {len(sentences)} evaluations."""

        add_log(f"📦 Batch {batch_idx+1}/{total_batches}: sentences {batch_start+1}-{batch_end}")
        _progress_store["progress"] = int((batch_idx / total_batches) * 100)

        for retry_attempt in range(MAX_RETRIES):
            try:
                add_log(f"  📡 Sending API request (attempt {retry_attempt+1})...")
                response = client.models.generate_content(
                    model=model, contents=prompt,
                    config=types.GenerateContentConfig(response_mime_type='application/json')
                )
                if response.text:
                    response_text = response.text.strip()
                    if response_text.startswith('```'):
                        response_text = re.sub(r'^```json?\s*', '', response_text)
                        response_text = re.sub(r'\s*```$', '', response_text)
                    result = json.loads(response_text)
                    evaluations = result.get("evaluations", [])
                    for eval_item in evaluations:
                        sent_id = eval_item.get("id")
                        score = eval_item.get("score", "")
                        reason = eval_item.get("reason", "")
                        if score == 3 or score == "3":
                            reason = ""
                        all_evaluations[sent_id] = {"score": score, "reason": reason}
                    add_log(f"  ✅ Evaluated {len(evaluations)} sentences")
                    batch_success = True
                    break
                else:
                    add_log(f"  ⚠️ Empty response, retrying...")
                    if retry_attempt < MAX_RETRIES - 1:
                        time.sleep(RETRY_DELAY * (2 ** retry_attempt) + random.uniform(0, 1))
            except Exception as e:
                add_log(f"  ⚠️ Retry {retry_attempt+1}: {str(e)[:80]}")
                if retry_attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY * (2 ** retry_attempt) + random.uniform(0, 1))

        if batch_end < total_sentences:
            rpm_limit = MODEL_RPM_LIMITS.get(model, 15)
            wait_time = max(BATCH_DELAY, 60.0 / rpm_limit)
            add_log(f"  ⏳ Waiting {wait_time:.0f}s before next batch...")
            time.sleep(wait_time)

    # Update alignment data
    for i, d in enumerate(alignment_data):
        sent_id = d["id"]
        if sent_id in all_evaluations:
            d["ai_score"] = all_evaluations[sent_id]["score"]
            d["ai_reason"] = all_evaluations[sent_id]["reason"]

    finish_progress(f"Evaluated {len(all_evaluations)} sentences")
    return jsonify({
        "alignment_data": alignment_data,
        "evaluated_count": len(all_evaluations)
    })


@app.route("/api/reevaluate", methods=["POST"])
def api_reevaluate():
    """Re-run local diff/WER calculation on (possibly edited) alignment data."""
    data = request.json
    alignment_data = data.get("alignment_data", [])
    if not alignment_data:
        return jsonify({"error": "No data to re-evaluate."}), 400

    total_subs = 0; total_dels = 0; total_ins = 0; total_refs = 0

    for d in alignment_data:
        true_text = d.get("true", "")
        asr_text = d.get("asr_nobreak", "") or d.get("asr", "")
        true_words = true_text.split()
        asr_words = asr_text.split()
        d["count"] = len(true_words)

        # Re-run diff
        result = run_alignment([true_text], [asr_text])
        if result["alignment_data"]:
            rd = result["alignment_data"][0]
            d["diffs"] = rd.get("diffs", [])
            d["wrong_count"] = rd.get("wrong_count", 0)
            d["wrong_list"] = rd.get("wrong_list", "")
            d["wer"] = rd.get("wer", 0)
            d["srr"] = rd.get("srr", "")
            d["score"] = rd.get("score", "")
            d["asr_displayed"] = rd.get("asr_displayed", asr_text)
            d["asr_separated"] = rd.get("asr_separated", asr_text)
            d["asr_nobreak"] = rd.get("asr_nobreak", asr_text)

        total_refs += len(true_words)
        stats = result.get("overall_stats", {})
        total_subs += stats.get("subs", 0)
        total_dels += stats.get("dels", 0)
        total_ins += stats.get("ins", 0)

    overall_wer = round((total_subs + total_dels + total_ins) / total_refs * 100, 1) if total_refs else 0
    overall_stats = {"subs": total_subs, "dels": total_dels, "ins": total_ins, "refs": total_refs}

    return jsonify({
        "alignment_data": alignment_data,
        "overall_wer": overall_wer,
        "overall_stats": overall_stats
    })


@app.route("/api/export-excel", methods=["POST"])
def api_export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font as OpenPyXLFont, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.text import InlineFont
        from openpyxl.cell.rich_text import CellRichText, TextBlock
    except ImportError:
        return jsonify({"error": "Install openpyxl: pip install openpyxl"}), 400

    data = request.json
    alignment_data = data.get("alignment_data", [])
    overall_wer = data.get("overall_wer", 0)
    overall_stats = data.get("overall_stats", {})

    if not alignment_data:
        return jsonify({"error": "No data to export."}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "ASR Evaluation"

    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = OpenPyXLFont(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        "Sentence No.", "Test Language", "Word Count", "ASR Result (Displayed)",
        "ASR Result (Separated)", "ASR Result (No Break)", "SRR",
        "Wrong Word Count", "Sentence Score", "AI Score", "AI Reason",
        "Differences", "Wrong Words", "Translation"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    column_widths = [6, 33, 11, 33, 33, 33, 6, 11, 11, 11, 25, 33, 33, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row_idx, d in enumerate(alignment_data, 2):
        plain_text_parts = [t for _, t in d.get("diffs", [])]
        plain_text_value = " ".join(plain_text_parts)

        row_data = [
            d.get("id", ""), d.get("true", ""), d.get("count", ""),
            d.get("asr_displayed", ""), d.get("asr_separated", ""),
            d.get("asr_nobreak", ""), d.get("srr", ""),
            d.get("wrong_count", ""), d.get("score", ""),
            d.get("ai_score", ""), d.get("ai_reason", ""),
            plain_text_value, d.get("wrong_list", ""), d.get("translation", "")
        ]
        ws.append(row_data)

        for cell in ws[row_idx]:
            cell.alignment = cell_alignment
            cell.border = thin_border

        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

        # Color the differences cell with per-word formatting (RichText)
        diff_cell = ws.cell(row=row_idx, column=12)
        diffs = d.get("diffs", [])
        if diffs:
            try:
                rich_parts = []
                for i, (color, word) in enumerate(diffs):
                    prefix = " " if i > 0 else ""
                    if color == "red":
                        rich_parts.append(TextBlock(InlineFont(b=True, color="C62828"), prefix + word))
                    elif color == "blue":
                        rich_parts.append(TextBlock(InlineFont(b=True, color="1565C0"), prefix + word))
                    else:
                        rich_parts.append(TextBlock(InlineFont(rFont='Calibri', sz=10), prefix + word))
                diff_cell.value = CellRichText(rich_parts)
            except Exception:
                # Fallback: plain text with single color
                has_error = any(c == "red" for c, _ in diffs)
                has_insertion = any(c == "blue" for c, _ in diffs)
                if has_error:
                    diff_cell.font = OpenPyXLFont(color="C62828", size=10, bold=True)
                elif has_insertion:
                    diff_cell.font = OpenPyXLFont(color="1565C0", size=10, bold=True)

    ws.freeze_panes = "A2"

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ASR_Evaluation_{timestamp}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True, download_name=filename)


@app.route("/api/export-json", methods=["POST"])
def api_export_json():
    data = request.json
    alignment_data = data.get("alignment_data", [])
    overall_wer = data.get("overall_wer", 0)

    if not alignment_data:
        return jsonify({"error": "No data to export."}), 400

    export_data = {
        "metadata": {
            "exported_at": datetime.datetime.now().isoformat(),
            "total_sentences": len(alignment_data),
            "overall_wer": overall_wer
        },
        "scoring_guide": {
            "3": "No wrong/missing/additional word. Perfect sentence.",
            "2.5": "Some errors but no problem understanding the sentence.",
            "2": "Some errors + grammar issues, but overall meaning is clear.",
            "1.5": "Partially able to guess and understand subject/details.",
            "1": "Unable to understand or no text transcribed."
        },
        "sentences": []
    }

    for d in alignment_data:
        export_data["sentences"].append({
            "id": d.get("id"),
            "true": d.get("true"),
            "asr": d.get("asr_nobreak"),
            "word_count": d.get("count"),
            "wer": d.get("wer", 0),
            "wrong_count": d.get("wrong_count"),
            "wrong_words": d.get("wrong_list", "").split(",") if d.get("wrong_list") else [],
            "ai_score": d.get("ai_score", ""),
            "ai_reason": d.get("ai_reason", ""),
            "translation": d.get("translation", "")
        })

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ASR_Evaluation_Input_{timestamp}.json"
    filepath = os.path.join(EXPORTS_DIR, filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, indent=2, ensure_ascii=False)

    return send_file(filepath, as_attachment=True, download_name=filename)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
